"""
export.py — generate a per-user Excel progress tracker (.xlsx) from the
applications table. Returns raw bytes so the web app can stream it as a download.

Requires: openpyxl   (pip install openpyxl)
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db

HEADERS = ["Company", "Job Title", "Location", "Status", "Applied On",
           "Notes", "Link"]

STATUS_FILL = {
    "interested": "FFF3CD",
    "applied":    "CCE5FF",
    "interview":  "D4EDDA",
    "offer":      "C3E6CB",
    "rejected":   "F8D7DA",
}


def build_tracker_xlsx(user_id: int, username: str = "") -> bytes:
    rows = db.list_applications(user_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    title_font = Font(bold=True, size=14)
    ws["A1"] = f"Job Hunt Tracker — {username}".strip(" —")
    ws["A1"].font = title_font
    ws.append([])  # spacer row

    header_row = 3
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    r = header_row + 1
    for a in rows:
        ws.cell(row=r, column=1, value=a["company"])
        ws.cell(row=r, column=2, value=a["title"])
        ws.cell(row=r, column=3, value=a["location"])
        status = (a["status"] or "interested").lower()
        sc = ws.cell(row=r, column=4, value=status.capitalize())
        if status in STATUS_FILL:
            sc.fill = PatternFill("solid", fgColor=STATUS_FILL[status])
        ws.cell(row=r, column=5, value=(a["applied_at"] or "")[:10])
        ws.cell(row=r, column=6, value=a["notes"] or "")
        link = ws.cell(row=r, column=7, value="Open")
        if a["url"]:
            link.hyperlink = a["url"]
            link.font = Font(color="2563EB", underline="single")
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=(c == 6))
        r += 1

    widths = [30, 40, 24, 14, 13, 40, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    data = build_tracker_xlsx(1, "demo")
    with open("tracker_demo.xlsx", "wb") as f:
        f.write(data)
    print(f"wrote tracker_demo.xlsx ({len(data)} bytes)")
