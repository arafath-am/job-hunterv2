#!/usr/bin/env python3
"""Generate Excel reports of tracked vs untracked companies."""
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

conn = sqlite3.connect("jobhunter.db")

# ── Styles ────────────────────────────────────────────────────
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill_green = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
header_fill_red = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
data_font = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, fill):
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

def style_data(ws, max_row):
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border

# ── Sheet 1: Tracking (456 companies) ────────────────────────
wb = Workbook()

ws1 = wb.active
ws1.title = "Collecting"
ws1.append(["#", "Company", "ATS Platform", "Cap-Exempt", "Sponsor", "Endpoint", "Active Jobs"])

rows = conn.execute("""
    SELECT c.brand, c.ats, c.cap_exempt, c.sponsor, c.endpoint,
           (SELECT COUNT(*) FROM jobs j WHERE j.company = c.company_name AND j.active = 1) as job_count
    FROM companies c
    WHERE c.endpoint IS NOT NULL AND c.ats IS NOT NULL
    ORDER BY c.ats, c.brand
""").fetchall()

for i, r in enumerate(rows, 1):
    ws1.append([i, r[0], r[1], r[2] or "", r[3] or "", r[4] or "", r[5]])

style_header(ws1, 1, header_fill_green)
style_data(ws1, len(rows) + 1)

# Auto-width
for col in ws1.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws1.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

# ── Sheet 2: Not Tracking (437 companies) ────────────────────
ws2 = wb.create_sheet("Not Collecting")
ws2.append(["#", "Company", "Status", "Cap-Exempt", "Sponsor", "ATS (if identified)", "Careers URL"])

rows2 = conn.execute("""
    SELECT brand, resolve_status, cap_exempt, sponsor, ats, careers_url
    FROM companies
    WHERE endpoint IS NULL OR ats IS NULL
    ORDER BY brand
""").fetchall()

for i, r in enumerate(rows2, 1):
    ws2.append([i, r[0], r[1] or "unknown", r[2] or "", r[3] or "", r[4] or "", r[5] or ""])

style_header(ws2, 1, header_fill_red)
style_data(ws2, len(rows2) + 1)

for col in ws2.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

# ── Sheet 3: Summary ─────────────────────────────────────────
ws3 = wb.create_sheet("Summary")
ws3.insert_rows(1)

summary_data = [
    ["Job Hunter v2 — Company Coverage Report"],
    [],
    ["Total Companies", 893],
    ["Actively Collecting", len(rows)],
    ["Not Collecting", len(rows2)],
    ["Active Job Postings", conn.execute("SELECT COUNT(*) FROM jobs WHERE active=1").fetchone()[0]],
    [],
    ["ATS Breakdown (Collecting)"],
]

ats_stats = conn.execute("""
    SELECT ats, COUNT(*) FROM companies 
    WHERE endpoint IS NOT NULL AND ats IS NOT NULL 
    GROUP BY ats ORDER BY COUNT(*) DESC
""").fetchall()

for ats, cnt in ats_stats:
    summary_data.append([f"  {ats}", cnt])

summary_data.extend([
    [],
    ["Not Collecting Breakdown"],
])

status_stats = conn.execute("""
    SELECT COALESCE(resolve_status, 'unknown'), COUNT(*) FROM companies 
    WHERE endpoint IS NULL OR ats IS NULL 
    GROUP BY resolve_status ORDER BY COUNT(*) DESC
""").fetchall()

for status, cnt in status_stats:
    summary_data.append([f"  {status}", cnt])

for row_data in summary_data:
    ws3.append(row_data)

ws3["A1"].font = Font(name="Arial", bold=True, size=14)
ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 15

# ── Save ──────────────────────────────────────────────────────
outpath = "company_coverage_report.xlsx"
wb.save(outpath)
conn.close()
print(f"Saved to {outpath}")
print(f"  Sheet 1: {len(rows)} collecting companies")
print(f"  Sheet 2: {len(rows2)} not collecting companies")
print(f"  Sheet 3: Summary")
